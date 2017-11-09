import os.path
import collections
import xlrd
import json
import xlsxwriter
import string
import datetime
from django.views.generic import View , TemplateView ,ListView
from django.views.generic.edit import CreateView, UpdateView, FormView
from django.contrib import messages
from django.core.exceptions import ObjectDoesNotExist
from django.db.models import Count, Q, Sum
from django.shortcuts import get_object_or_404
from django.http import JsonResponse
from django.http import HttpResponseRedirect, HttpResponse
from django.shortcuts import render
from .forms import *
from MyANSRSource.models import ProjectTeamMember, ProjectDetail
from django.forms.formsets import formset_factory
from django.core.urlresolvers import reverse, reverse_lazy
from django.db import transaction
from shutil import copyfile
from MyANSRSource.reportviews import generateDownload
from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.datavalidation import DataValidation
import logging
logger = logging.getLogger('MyANSRSource')

s1_penalty_count = 0.5
s2_penalty_count = 0.3
s3_penalty_count = 0.2

try:
    dict.iteritems
except AttributeError:
    # Python 3
    def itervalues(d):
        return iter(d.values())

    def iteritems(d):
        return iter(d.items())
else:
    # Python 2
    def itervalues(d):
        return d.itervalues()

    def iteritems(d):
        return d.iteritems()


class ChooseTabs(FormView):
    template_name = 'reviewreport_create_form_1.html'
    form_class = ChooseMandatoryTabsForm
    success_url = reverse_lazy('choose_tabs')

    def form_valid(self, form):
        user_tab = {}
        order_number = {}
        chapter_list = self.request.POST.getlist("to[]")
        users = {k: v for k, v in self.request.POST.items() if k.startswith('user_')}
        order = {k: v for k, v in self.request.POST.items() if k.startswith('order_')}
        # print users
        for k, v in users.iteritems():
            tab_id = k.split('_')
            user_tab[tab_id[1]] = v

        for k, v in order.iteritems():
            order_id = k.split('_')
            # print v
            if v:
                v = int(v)
            if v and v != 0:
                order_number[order_id[1]] = v
            else:
                # print 'else'
                messages.error(self.request, 'Order Number Cannot Be 0')
                return super(ChooseTabs, self).form_valid(form)

        # {u'user_3': u'256', u'user_1': u'255'}
        # print users
        # print order

        try:
            ProjectTemplateProcessModel.objects.get_or_create(template=form.cleaned_data['template'],
                                                              project=form.cleaned_data['project'],
                                                              qms_process_model=form.cleaned_data['qms_process_model'],
                                                              defaults={
                                                                  'created_by': self.request.user},)
        except Exception as e:
            # print "ChooseTabs", (str(e))
            logger.error(" {0} ".format(str(e)))

        cm_obj, chapter_component = ChapterComponent.objects.get_or_create(chapter=form.cleaned_data['chapter'],
                                                                           component=form.cleaned_data['component'],
                                                                           defaults={
                                                                               'created_by': self.request.user}, )
        # print form.cleaned_data['chapter'] , form.cleaned_data['component']
        # print"res", cm_obj
        # print "obj", chapter_component
        # print user_tab
        for k, v in user_tab.iteritems():
            # print form.cleaned_data['author']
            # k = int(k)
            # print k, order_number[k]
            obj, created = QASheetHeader.objects.update_or_create(project=form.cleaned_data['project'],
                                                                  chapter=form.cleaned_data['chapter'],
                                                                  chapter_component=cm_obj,
                                                                  review_group_id=int(k),
                                                                  defaults={'author': form.cleaned_data['author'],
                                                                            'reviewed_by_id': int(v),
                                                                            'created_by': self.request.user,
                                                                            'order_number': order_number[k]}, )
            # print obj, created
            if not created:
                obj.updated_by = self.request.user
                obj.save()
        qa_obj = QASheetHeader.objects.filter(project=form.cleaned_data['project'], chapter_component=cm_obj).\
            values('review_group_id', 'author', 'reviewed_by_id', 'order_number')
        if chapter_list:
            chapter_list = set(chapter_list)
            for obj in qa_obj:
                for chapter in chapter_list:
                    cm_sub_obj, chapter_component = ChapterComponent.objects.get_or_create(chapter_id=chapter,
                                                                                           component=form.cleaned_data[
                                                                                            'component'],
                                                                                           defaults={
                                                                                             'created_by':
                                                                                             self.request.user}, )

                    qa_sub_obj, created = QASheetHeader.objects.update_or_create(project=form.cleaned_data['project'],
                                                                                 chapter_id=chapter,
                                                                                 chapter_component=cm_sub_obj,
                                                                                 review_group_id=obj['review_group_id'],
                                                                                 defaults={
                                                                                  'author_id': obj['author'],
                                                                                  'reviewed_by_id':
                                                                                  obj['reviewed_by_id'],
                                                                                  'created_by': self.request.user,
                                                                                  'order_number': obj['order_number']},
                                                                                 )
                    if not created:
                        qa_sub_obj.updated_by = self.request.user
                        qa_sub_obj.save()

        messages.info(self.request, "Successfully Saved")
        return super(ChooseTabs, self).form_valid(form)


def get_review(obj):
    try:
        s = ReviewReport.objects.filter(QA_sheet_header=obj.id, is_active=True). \
            values('id', 'review_item', 'defect', 'defect_severity_level__severity_type',
                   'defect_severity_level__severity_level', 'defect_severity_level__defect_classification',
                   'is_fixed', 'fixed_by__username', 'remarks', 'screen_shot', 'instruction')
    except Exception as e:
        s = None
        logger.error(" {0} ".format(str(e)))
    return s


def qa_sheet_header_obj(project, chapter, author, component=None, active_tab=None):
    # qa_sheet_header_obj(project, chapter, author=author)
    try:
        # print chapter, "chapter"
        # print active_tab, "-<active_tab->", project, chapter, author
        result = None
        if active_tab and component is not None:
            chapter_component_obj = ChapterComponent.objects.get(chapter=chapter, component=component)
            if active_tab == 'lambda':
                review_obj = QASheetHeader.objects.filter(project=project, chapter_component=chapter_component_obj,
                                                          author=author).values_list('review_group').\
                    order_by('order_number').first()
                review_obj = review_obj[0]
            else:
                review_obj = ReviewGroup.objects.get(id=active_tab)
            try:
                result = QASheetHeader.objects.get(project=project, chapter_component=chapter_component_obj,
                                                   author=author,
                                                   review_group=review_obj)
            except Exception as e:
                logger.error(" {0} ".format(str(e)))

        else:
            result = QASheetHeader.objects.filter(project=project, chapter=chapter, author=author)

    except ObjectDoesNotExist as e:
        # print "except"
        logger.error(" {0} ".format(str(e)))
    return result


def get_review_group(project=None, chapter=None, is_author=False, component=False):
    # print "get_review_group"
    obj = ReviewGroup.objects.all()
    # print is_author , project,chapter,component
    try:
        if not is_author and project and chapter:
            obj = QASheetHeader.objects.filter\
                (project=project, chapter_component=
                 ChapterComponent.objects.get(chapter=chapter, component=component)).\
                values('review_group_id', 'review_group__name', 'review_group__alias').order_by('order_number')

            # for s in obj:
                #  print s.review_group_id,s.review_group__name,s.review_group__alias
    except Exception as e:
        logger.error(" {0} ".format(str(e)))
        pass
    return obj


def mark_as_completed(request):
    result = False
    can_show_button = False

    if request.GET.get('is_lead') == 'true':
        # print "in if true"
        try:
            QASheetHeader.objects.filter(project=request.GET.get('project_id')).update(review_group_status=True,
                                                                                       author_feedback_status=True)
            ProjectTemplateProcessModel.objects.filter(project=request.GET.get('project_id')).\
                update(lead_review_status=True, lead_review_feedback=request.GET.get('feedback'))
            result = True
        except Exception as e:
            # print str(e)
            logger.error(" {0} ".format(str(e)))

    else:# print request.session['c_project'], request.session['c_chapter'],request.session['c_component'],request.GET.get('tab_id')
        # print "in else"
        try:
            QASheetHeader.objects.filter(project=request.session['c_project'],
                                         chapter_component=ChapterComponent.objects.
                                         get(chapter=request.session['c_chapter'],
                                             component=request.session['c_component']),
                                         review_group=request.GET.get('tab_id')).update(review_group_status=True,
                                                                                        author_feedback_status=True)
            result = True
            can_show_button = QASheetHeader.objects.filter((Q(review_group_status=False) |
                                                            Q(author_feedback_status=False)),
                                                           project=request.session['c_project']).exists()
            # print can_show_button
            if not can_show_button:
                obj = ProjectTemplateProcessModel.objects.get(project=request.session['c_project'])
                if obj.lead_review_status is False:
                    can_show_button = True
        except Exception as e:
            # print str(e)
            logger.error(" {0} ".format(str(e)))
    data = {"result": result, "can_show_button": can_show_button}
    return HttpResponse(
        json.dumps(data),
        content_type="application/json"
    )


def get_template_process_review(request):
    # print request.GET
    request.session['c_project'] = project = request.GET.get('project_id')
    template = request.GET.get('template_id')
    qms_process_model = request.GET.get('qms_process_model')
    request.session['c_chapter'] = chapter = request.GET.get('chapter')
    author = request.GET.get('author')
    request.session['c_component'] = component = request.GET.get('component')
    tabs = {}
    tab_name = {}
    team_members = {}
    user_tab = {}
    tab_order = {}
    can_edit = {}
    exclude_list = []
    config_missing = False
    show_lead_complete = False
    is_completed = False
    try:
        # print template,qms_process_model
        is_completed = ProjectTemplateProcessModel.objects.filter(project=project, lead_review_status=True).exists()
        obj = TemplateProcessReview.objects.filter(template=template, qms_process_model=qms_process_model). \
            order_by('id')
        # print "obj",obj,obj.query
        if not obj:
            config_missing = True
        members_obj = ProjectTeamMember.objects.filter(project=project, member__is_active=True)
        # print members_obj.query
        # for s in members_obj:
        # print "mem", s
        qa = QASheetHeader.objects.filter(project=project).values_list('chapter_component', flat=True)
        existing_chapters = ChapterComponent.objects.filter(id__in=qa, component=component).values_list("chapter",
                                                                                                        flat=True)
        cm_obj, created = ChapterComponent.objects.get_or_create(chapter_id=chapter, component_id=component, defaults={
                                                                                             'created_by':
                                                                                              request.user}, )
        # Chapter.objects.
        try:
            # qa_obj = qa_sheet_header_obj(project, chapter, author=author)
            qa_obj = QASheetHeader.objects.filter(project=project, chapter_component=cm_obj)
            # print "qa_obj" ,qa_obj
            # qa_obj.filter(chapter_component=cm_obj)
        except Exception as e:
            pass
            # print "exception", str(e)
            qa_obj_count = 0
        # print "qa_obj" ,qa_obj
        qa_obj_count = qa_obj.count()
        for members in members_obj:
            if int(members.member_id) != int(author):
                team_members[int(members.member_id)] = str(members.member.username)

        for ele in obj:
            tabs[str(ele.review_group)] = bool(ele.is_mandatory)
            tab_name[str(ele.review_group)] = int(ele.review_group.id)
            if qa_obj_count > 0:
                try:
                    tab_user = qa_obj.get(review_group=ele.review_group)
                    review_report_obj = ReviewReport.objects.filter(QA_sheet_header__in=qa_obj,
                                                                    QA_sheet_header__review_group=ele.review_group)
                    if review_report_obj:
                        can_edit[str(ele.review_group.id)] = False
                    else:
                        if tab_user.review_group_status and tab_user.author_feedback_status:

                            can_edit[str(ele.review_group.id)] = False
                        else:
                            can_edit[str(ele.review_group.id)] = True
                except ObjectDoesNotExist:
                    tab_user = None
                if tab_user is not None:
                    user_tab[str(ele.review_group)] = int(tab_user.reviewed_by.id)
                    tab_order[str(ele.review_group)] = int(tab_user.order_number)
                else:
                    user_tab[str(ele.review_group)] = None
                    tab_order[str(ele.review_group)] = None
        if qa_obj_count > 0 and qa_obj_count == len(exclude_list):
            show_lead_complete = True

    except Exception as e:
        logger.error(" {0} ".format(str(e)))
        tabs = team_members = tab_name = ''
        config_missing = True
    if is_completed:  # mark all tabs as  disabled when qms for the project is marked completed
        can_edit = {x: False for x in can_edit}
    exclude_list = [k for k, v in can_edit.iteritems() if v is False]
    current_tab = False

    if exclude_list:
        try:
            tmp_obj = qa_obj.filter(review_group__in=exclude_list).exclude(Q(review_group_status=True) &
                                                                               Q(author_feedback_status=True)).\
                values_list("review_group_id", flat=True).first()
            # print "if current_tab", current_tab
            if tmp_obj:
                current_tab = int(tmp_obj)
            else:
                # print "else"
                tmp_obj = qa_obj.filter(Q(review_group_status=True) & Q(author_feedback_status=True)).\
                    order_by('-order_number').values_list("order_number", flat=True).first()
                # print tmp_obj
                # print "else tmp_obj tab", tmp_obj
                current_tab = qa_obj.filter(order_number=int(tmp_obj) + 1).values_list("review_group_id", flat=True).first()
                current_tab = int(current_tab)
        except Exception as e:
            logger.error(" {0} ".format(str(e)))
            # print str(e)
    project_obj = Project.objects.get(id=int(project))
    chapter = Chapter.objects.filter(book=project_obj.book).exclude(id__in=existing_chapters).values_list('id', 'name')
    chapter = dict((x, y) for x, y in chapter)

    context_data = {'tabs': tabs, 'tab_name': tab_name, 'team_members': team_members, 'user_tab': user_tab,
                    'tab_order': tab_order, 'config_missing': config_missing, "can_edit": can_edit,
                    "current_tab": current_tab, "show_lead_complete": show_lead_complete, "chapters": chapter,
                    "is_completed": is_completed}
    # print context_data
    return HttpResponse(
        json.dumps(context_data),
        content_type="application/json"
    )


def forbidden_access(self, form, project, message_code, chapter=None):
    msg_dict = {'not_assigned': "Sorry You are not assigned to  this chapter",
                "config_missing": "Sorry configuration is missing please contact your manager",
                "wait": "Sorry You cant access this chapter till review is completed",
                "config_missing_manager": "Hey you didn't configure for this review please do it",
                "previous_tab_wait": " You cannot access this chapter till previous review is completed",
                "previous_tab_wait_pm": " Please wait  till previous review is completed",
                }
    if message_code == "previous_tab_wait_pm":
        result = False
    else:
        result = True

    messages.error(self.request, msg_dict[message_code])
    try:
        c = get_review_group(project, chapter,  component=self.request.session['component'])
    except Exception as e:
        pass
        # print str(e)

    return render(self.request, self.template_name, {'form': form,
                                                     'review_group': get_review_group(project, chapter,
                                                                                      component=self.request.session['component']),
                                                     "need_button": result})


def get_work_book(qms_form, reports, obj):
    qms_data_list, severity_count = initial_to_form(reports)
    qms_formset = formset_factory(
        qms_form, max_num=1, can_delete=True
    )
    if len(qms_data_list):
        qms_formset = qms_formset(initial=qms_data_list)
    else:
        qms_formset = formset_factory(
            qms_form, extra=1, max_num=1, can_delete=True
        )

    score = {}
    tmp_weight = {}
    defect_density = {}
    sev_count = []
    # print severity_count
    s = SeverityLevelMaster.objects.filter(is_active=True)
    for k, v in severity_count.iteritems():
        # print "k,v", k,v
        severity_level_obj = s.get(id=int(k))
        # print severity_level_obj
        if severity_level_obj.name.lower() == "s0":
            score[k] = 0
            defect_density[k] = 0
            tmp_weight[k] = 0
        else:
            sev_count.append(v)
            tmp_weight[k] = float(severity_level_obj.penalty_count) * v
            score[k] = 100 - (tmp_weight[k])
            if obj.count > 0:
                defect_density[k] = round(((tmp_weight[k] / obj.count) * 100), 2)
            else:
                defect_density[k] = 0
    # print list(severity_count.itervalues())
    # for k,v in tmp_weight.iteritems():
    #     print k,v
    total_count = sum(sev_count)
    weight = sum(tmp_weight.itervalues())
    if weight != 0:
        total_score = 100 - sum(tmp_weight.itervalues())
    else:
        total_score = 0
    total_defect_density = sum(defect_density.itervalues())

    # below to pack multiple variables in named tuple
    result = collections.namedtuple('result', ['a', 'b', 'c', 'd', 'e', 'f', 'g'])
    resultant_obj = result(severity_count, score, total_score, total_count, defect_density, total_defect_density,
                           qms_formset)
    return resultant_obj  # resultant_obj


def initial_to_form(reports):
    qms_data = {}
    qms_data_list = []
    severity_count = {}
    try:
        if reports and len(reports) != 0:
            for eachData in reports:

                for k, v in eachData.iteritems():
                    qms_data[k] = v
                    if k == 'id':
                        r_obj = ReviewReport.objects.get(id=int(v))
                        if r_obj.screen_shot:
                            qms_data['screen_shot_url'] = r_obj.screen_shot.url
                            # print "url",  r_obj.screen_shot.path
                        else:
                            qms_data['screen_shot_url'] = None
                        qms_data['qms_id'] = v
                    if k == 'review_item':
                        qms_data['review_item'] = v

                    if k == 'defect':
                        qms_data['defect'] = v

                    if k == 'instruction':
                        qms_data['instruction'] = v

                    if k == 'defect_severity_level__severity_type':
                        qms_data['severity_type'] = v

                    if k == 'defect_severity_level__severity_level':
                        qms_data['severity_level'] = v
                        if v in severity_count:
                            v = int(v)

                            severity_count[v] += 1
                        else:
                            severity_count[v] = 1

                    if k == 'defect_severity_level__defect_classification':
                        qms_data['defect_classification'] = v

                    if k == 'is_fixed':
                        qms_data['is_fixed'] = v

                    if k == 'screen_shot':
                        # url = ReviewReport.objects.get(id=obj.id)
                        qms_data['screen_shot'] = v
                        # if v:
                        #     qms_data['screen_shot_url'] = v

                    if k == 'fixed_by__username':
                        qms_data['fixed_by'] = v

                    if k == 'remarks':
                        qms_data['remarks'] = v
                        # qms_data['clear_screen_shot'] = False
                # print "" qms_data
                qms_data_list.append(qms_data.copy())
                qms_data.clear()
    except Exception as e:
        pass
        # print str(e)
    # print qms_data_list
    return qms_data_list, severity_count


class AssessmentView(TemplateView):
    template_name = "ansrS_QA_Tmplt_Assessment (Non Platform) QA sheet_3.3.html"

    def get_context_data(self, **kwargs):
        context = super(AssessmentView, self).get_context_data(**kwargs)
        # print "user", self.request.user
        form = BaseAssessmentTemplateForm()
        context['form'] = form
        # context['review_group'] = get_review_group()

        return context

    def post(self, request):
        form = BaseAssessmentTemplateForm(request.POST)
        request.session['hide_export'] = False
        # for sa in request.POST :
        # print request.POST
        # reports = template_id = None
        active_tab = request.POST.get('active_tab')
        request.session['active_tab'] = request.POST.get('active_tab')
        # print "after ass " , active_tab
        # print request.POST
        if form.is_valid():
            project = form.cleaned_data['project']
            chapter = form.cleaned_data['chapter']
            author = form.cleaned_data['author']
            component = form.cleaned_data['component']
            BaseAssessmentTemplateForm(initial={'project': project, 'chapter': chapter, 'author': author,
                                                'component': component })
            if request.user is not author:
                get_review_group(project, chapter, is_author=False, component=component)

            try:
                request.session['project'] = project
                request.session['chapter'] = chapter
                request.session['author'] = author
                request.session['component'] = component
                request.session['chapter_component'] = ChapterComponent.objects.get(chapter=chapter, component=component)

                ptpm_obj = ProjectTemplateProcessModel.objects.get(project=project)
                # if int(ptpm_obj.qms_process_model.product_type) == 1:
                #     request.session['hide_export'] = True
                request.session['template_id'] = ptpm_obj.template_id
                obj = qa_sheet_header_obj(project, chapter, author, component, active_tab)
                # print "status", obj.author_feedback_status
                is_pm = ProjectDetail.objects.filter(Q(deliveryManager=request.user)|Q(pmDelegate=request.user),
                                                     project=project).exists()
                request.session['is_pm'] = is_pm
                # print obj.order_number
                #  previous tab completion check
                if obj.order_number != 1:
                    order_number = int(obj.order_number)-1
                    # print "order_number", order_number
                    prev_tab_obj = QASheetHeader.objects.filter(project=project,
                                                                chapter_component=request.session['chapter_component'],
                                                                order_number=order_number).first()
                    if is_pm:
                        # print "first if is pm"
                        # print prev_tab_obj.review_group_status , prev_tab_obj.author_feedback_status
                        if prev_tab_obj.review_group_status is True and \
                                        prev_tab_obj.author_feedback_status is True:
                            pass
                        else:
                            # print "is pm, else"
                            return forbidden_access(self, form, project, "previous_tab_wait_pm", chapter)
                    if not is_pm:
                        if prev_tab_obj.review_group_status and prev_tab_obj.author_feedback_status:
                            pass
                        else:
                            return forbidden_access(self, form, project, "previous_tab_wait", chapter)

                if request.user == obj.reviewed_by:
                    request.session['reviewer_logged_in'] = True
                    # for s in obj:
                    #     print s
                    if obj.review_group_status is True and obj.author_feedback_status is False:
                        messages.info(self.request, " Please wait till author submit their feedback")
                else:
                    request.session['reviewer_logged_in'] = False

                if obj is None:
                    if is_pm:
                        msg = "config_missing_manager"
                    else:
                        msg = "config_missing"
                    return forbidden_access(self, form, project, msg, chapter)
                else:
                    # print "in else"
                    can_access = 0
                    qms_team_members = [obj.reviewed_by, author]
                    if not is_pm:
                        if request.user == author:
                            request.session['author_logged_in'] = True
                            if not obj.review_group_status and not obj.author_feedback_status:
                                return forbidden_access(self, form, project, "wait", chapter)
                        else:
                            # print request.user , author
                            request.session['author_logged_in'] = False
                    else:
                        request.session['author_logged_in'] = False
                    if not is_pm and request.user not in qms_team_members:
                        request.session['hide_export'] = True
                        # return forbidden_access(self, form, project, "not_assigned", chapter)

                project_template_process_model_obj = ProjectTemplateProcessModel.objects.get(project=project)
                template_id = request.session['template_id'] = project_template_process_model_obj.template_id
                request.session['QA_sheet_header_id'] = obj.id
                qms_form = review_report_base(template_id, project, ChapterComponent.objects.get(chapter=chapter,
                                                                                                 component=component),
                                              request_obj=self.request, tab=obj.review_group_id)

                return render_common(obj, qms_form, request)

            except Exception as e:
                messages.info(request, "No data filled against that project")
                logger.error(" {0} ".format(str(e)))
                return render(request, "ansrS_QA_Tmplt_Assessment (Non Platform) QA sheet_3.3.html",
                              {'form': BaseAssessmentTemplateForm(request.POST), })

        else:
            # print form.errors()
            return render(request, "ansrS_QA_Tmplt_Assessment (Non Platform) QA sheet_3.3.html",
                          {'form': BaseAssessmentTemplateForm(request.POST), })


class ReviewReportManipulationView(AssessmentView):
    @transaction.atomic
    def post(self, request):
        fail = 0
        qms_data = {}
        qms_data_list = []
        request.session['active_tab'] = active_tab = request.POST.get('active_tab1')
        qms_form = review_report_base(request.session['template_id'], request.session['project'],
                                      request_obj=self.request, tab=active_tab)
        qms_formset = formset_factory(
            qms_form,  max_num=1, can_delete=True
        )
        AllowedFileTypes = ['jpg', 'png', 'pdf', 'xlsx', 'xls', 'docx', 'doc', 'jpeg', 'eml', 'zip', 'gz', '7z']

        forbidden_file_type = False

        q_form = qms_formset(request.POST, request.FILES)
        # print request.POST
        # print q_form.is_valid()
        # print q_form.errors
        forbidden_file = False
        if 'import_file' in request.FILES and request.FILES['import_file']:
            if self.request.FILES['import_file'].name.split(".")[-1] == "xlsx":
                return import_review(self.request, self.request.FILES['import_file'])
            else:
                forbidden_file = True

        if q_form.is_valid() and not forbidden_file:

            for form_elements in q_form:
                # print form_elements
                if form_elements.cleaned_data['DELETE'] is True:
                    ReviewReport.objects.filter(id=form_elements.cleaned_data['qms_id']).update(is_active=False)
                else:
                    del(form_elements.cleaned_data['DELETE'])
                    for k, v in form_elements.cleaned_data.iteritems():
                        qms_data[k] = v
                    qms_data_list.append(qms_data.copy())
                    qms_data.clear()
            for obj in qms_data_list:
                if obj['severity_type'] == "":
                    continue
                if obj['qms_id'] > 0:
                    report = ReviewReport.objects.get(id=obj['qms_id'])
                    # print obj['qms_id']
                else:
                    # print obj['qms_id']
                    report = ReviewReport()
                    report.created_by = request.user
                if not request.session['author_logged_in']:
                    report.review_item = obj['review_item']
                    report.defect = obj['defect']
                    if 'instruction' in obj:
                        report.instruction = obj['instruction']
                    if obj['clear_screen_shot']:
                        report.screen_shot = None
                    else:

                        if 'screen_shot' in obj and obj['screen_shot']:
                            # extension = os.path.splitext(obj['screen_shot'])[1]
                            # if request.FILES['admin_action_attachment'].name.split(".")[-1] not in AllowedFileTypes:
                            if obj['screen_shot'].name.split(".")[-1] not in AllowedFileTypes:

                                messages.error(request, "You can't upload this file type")

                                forbidden_file_type = True

                            report.screen_shot = obj['screen_shot']
                report.is_fixed = obj['is_fixed']

                report.remarks = obj['remarks']
                if len(obj['is_fixed']) > 0:
                    report.fixed_by = request.user
                else:
                    report.fixed_by = None

                try:
                    defect_obj = DefectSeverityLevel.objects.filter(
                                                                    severity_type=obj['severity_type'],
                                                                    ).first()

                    report.defect_severity_level = defect_obj
                    qa_obj = QASheetHeader.objects.get(id=request.session['QA_sheet_header_id'])
                    report.QA_sheet_header = qa_obj
                    report.updated_by = request.user
                    report.save()
                    try:
                        qa_obj.count = int(request.POST.get('questions'))
                        qa_obj.save()
                    except Exception as e:
                        logger.error(" {0} ".format(str(e)))

                except DefectSeverityLevel.DoesNotExist as e:
                    logger.error(" {0} ".format(str(e)))
                    fail += 1

        else:
            logger.error(" {0} ".format(str(q_form.errors)))
            if forbidden_file:
                messages.error(self.request, "please upload only xlsx file")
            else:
                messages.error(request, "Sorry please try again")
            qa_obj = QASheetHeader.objects.get(id=request.session['QA_sheet_header_id'])
            return HttpResponseRedirect(
                reverse(u'review_redirect_view', kwargs={'id': request.session['QA_sheet_header_id'],
                                                         'chapter_component_id': qa_obj.chapter_component.id,
                                                         'review_group_id': qa_obj.review_group.id}))

        if fail == 0:
            if forbidden_file_type:
                messages.error(request, "You can't upload this file type but your data is saved")
            else:
                messages.info(request, "successfully saved")
        else:
            messages.error(request, "Configuration is Missing")

        obj = qa_sheet_header_obj(request.session['project'], request.session['chapter'], request.session['author'],
                                  request.session['component'], active_tab)
        return render_common(obj, qms_form, self.request)


def render_common(obj, qms_form, request):
    reports = get_review(obj)
    result = get_work_book(qms_form, reports, obj)
    severity_level_obj = SeverityLevelMaster.objects.filter(is_active=True).values_list('name', 'id').order_by('name')
    # . #  exclude(name__icontains='S0')

    try:
        form = BaseAssessmentTemplateForm(
            initial={'project': request.session['project'], 'chapter': request.session['chapter'],
                     'author': request.session['author'], 'component': request.session['component']})
    except:
        form = BaseAssessmentTemplateForm()
    # messages.success(request, "successfully saved")
    s = get_review_group(request.session['project'], request.session['chapter'], component=request.session['component'])

    ptpm_obj = ProjectTemplateProcessModel.objects.get(project=obj.project)
    # print s
    return render(request, "ansrS_QA_Tmplt_Assessment (Non Platform) QA sheet_3.3.html",
                  {'form': form, 'defect_master': DefectTypeMaster.objects.all(),
                   'reports': reports, 'review_formset': result[6], "author_feedback_status":
                       obj.author_feedback_status, "reviewer_feedback_status": obj.review_group_status,
                   "reviewer_feedback": obj.review_group_feedback, "author_feedback": obj.author_feedback,
                   "lead_review_feedback": ptpm_obj.lead_review_feedback, 'template_id': request.session['template_id'],
                   'review_group': s, 'questions': obj.count, 'severity_count': result[0],
                   'project': request.session['project'], 'score': result[1], 'total_score': result[2],
                   'total_count': result[3], 'defect_density': result[4], 'total_defect_density': result[5],
                   'severity_level': severity_level_obj, "need_button": True})


def fetch_severity(request):
    # print request.GET
    template_id = request.GET.get('template_id')
    project_id = request.GET.get('project_id')
    severity = request.GET.get('severity_type')
    request.session['active_tab'] = request.GET.get('active_tab')
    review_group = ReviewGroup.objects.get(id=request.GET.get('active_tab'))
    severity_classification = request.GET.get('severity_classification')
    # print request.GET
    obj = ProjectTemplateProcessModel.objects.get(template_id=template_id, project_id=project_id)

    try:
        if int(obj.qms_process_model.product_type) == 1:
            "if part"
            media_team = True
        else:
            "else part"
            media_team = False
    except Exception as e:
        pass
        # print str(e)

    if not media_team:
        try:
            # print "severity", severity
            obj = DefectSeverityLevel.objects.filter(severity_type=severity).first()
            # print "obj" , obj
            # logger.error("query {0} ".format(obj.query))
            context_data = {'severity_level': str(obj.severity_level), 'defect_classification':
                { str(obj.defect_classification): obj.defect_classification.id}}
        except IndexError as e:
            logger.error("query {0} ".format(str(e)))
            context_data = {'configuration_missing': True}

    else:
        if request.GET.get('triggered_by') == 'severity_type':
            # print"im in buddy"
            try:
                s = DefectSeverityLevel.objects.filter(severity_type=severity).values_list("defect_classification__name",
                                                                                           "defect_classification__id")
                classification_dict = dict((str(x), int(y)) for x, y in s)
                context_data = {'classification_dict': classification_dict, "triggered_by": request.GET.get('triggered_by')}

            except:
                context_data = {'configuration_missing': True, "triggered_by": request.GET.get('triggered_by')}
            return HttpResponse(
                json.dumps(context_data),
                content_type="application/json"
            )
        if severity_classification is not None:
            is_media = True
        else:
            is_media = False
        try:
            if is_media:
                try:
                    obj = DefectSeverityLevel.objects.filter(severity_type_id=severity,
                                                             defect_classification_id=severity_classification).first()
                except Exception as e:
                    pass
                    # print str(e)

            else:
                obj = DefectSeverityLevel.objects.filter(severity_type_id=severity).first()
            # logger.error("query {0} ".format(obj.query))
            context_data = {'severity_level': str(obj.severity_level), 'defect_classification':
                {str(obj.defect_classification): obj.defect_classification.id}, "is_media": is_media}
        except IndexError as e:
            logger.error("query {0} ".format(str(e)))
            context_data = {'configuration_missing': True, "is_media": is_media}
    return HttpResponse(
            json.dumps(context_data),
            content_type="application/json"
        )


def fetch_members(project):
    user = ProjectTeamMember.objects.filter(project=project, member__is_active=True)
    return user


def fetch_author(request):
    project_id = request.GET.get('project_id')
    chapter_id = request.GET.get('chapter_id')
    component_id = request.GET.get('component_id')

    try:
        chapter_component = ChapterComponent.objects.get(chapter=chapter_id, component=component_id)
        author = QASheetHeader.objects.filter(project_id=project_id,
                                              chapter_component=chapter_component).values_list('author', flat=True).first()
    except Exception as e:
        logger.error("fetch_author{0}", str(e))
        author = None

    team_members = {}
    team = fetch_members(project_id)
    for members in team:
        team_members[int(members.member_id)] = str(members.member.username)
    context_data = {'author': str(author), 'team_members': team_members}
    return HttpResponse(
        json.dumps(context_data),
        content_type="application/json"
    )


def c_get_severity_count(project, name, template_id):
    s = 0
    try:
        if name.lower() == "s0":
            return s
        severity_level_obj = SeverityLevelMaster.objects.get(name__icontains=name)
        review_report_obj = ReviewReport.objects.filter(QA_sheet_header__in=QASheetHeader.objects.filter(
            project=project).values_list('id', flat=True),
                                                        defect_severity_level__severity_level=severity_level_obj). \
            values('id', 'defect_severity_level__severity_level__name').exclude(is_active=False). \
            annotate(s_count=Count('defect_severity_level'))

        for v in review_report_obj:
            for key, value in v.iteritems():
                if key is 's_count':
                    s += value
    except Exception as e:
        logger.error("qms format{0}", str(e))
    return s


def c_get_question_count(project):
    s = 0
    try:
        qa_object = QASheetHeader.objects.filter(project=project).values_list('count', flat=True)
        for v in qa_object:
            s += v
    except Exception as e:
        logger.error("qms format{0}", str(e))
    return s


def c_get_project_status(project):
    can_show_button = QASheetHeader.objects.filter(
        (Q(review_group_status=False) | Q(author_feedback_status=False)),

        project=project).exists()
    obj = None
    if not can_show_button:
        obj = ProjectTemplateProcessModel.objects.get(project=project)
        if obj.lead_review_status is False:
            can_show_button = True
    chapter_count = Chapter.objects.filter(book__project=project).count()
    qa_chapter_count = QASheetHeader.objects.filter(project=project).values('chapter').distinct().count()
    if chapter_count != qa_chapter_count:
        difference = chapter_count - qa_chapter_count
    else:
        difference = 0
    if obj is not None and obj.lead_review_status is True:
        difference = 0
    return can_show_button, difference


def c_get_defect_density(s1, s2, s3, q_count):
    # print s1, s2, s3, q_count
    if q_count == 0:
        return 0
    else:
        s1_dd = (s1 * s1_penalty_count)/q_count
        s2_dd = (s2 * s2_penalty_count)/q_count
        s3_dd = (s3 * s3_penalty_count)/q_count
        return str(round(((s1_dd + s2_dd + s3_dd)*100), 2))


class DashboardView(ListView):
    model = ReviewReport
    template_name = 'qms_dashboard.html'

    def get_context_data(self, **kwargs):
        context = super(DashboardView, self).get_context_data(**kwargs)
        context['projects'] = ProjectTemplateProcessModel.objects.filter(project__endDate__gte=datetime.date.today(),
                                                                         project__closed=False, project__active=True,
                                                                         project__in=ProjectDetail.objects.filter(
                    Q(deliveryManager=self.request.user) | Q(pmDelegate=self.request.user)).
                                                                         values('project')).\
            values('id', 'project', 'project_id', 'project__projectId', 'project__name', 'template_id',
                   'lead_review_status').\
            annotate(chapter_count=Count('project__book__chapter'))
        return context

    def post(self, request, *args, **kwargs):
        workbook = xlsxwriter.Workbook('qms_project_summary.xlsx')
        worksheet = workbook.add_worksheet()
        header = ['Project Id', 'Project', 'Manager', 'Chapters', 'S1', 'S2', 'S3', 'Questions',
                  'Defect Density', 'Status']
        header_length = len(header)
        header_column = list(string.ascii_uppercase)[:header_length]
        header_column = [s+"1" for s in header_column]
        header = zip(header_column, header)

        project_details = ProjectTemplateProcessModel.objects.filter(project__endDate__gte=datetime.date.today(),
                                                                     project__active=True,
                                                                     project__closed=False,
                                                                     project__in=ProjectDetail.objects.filter(
                    Q(deliveryManager=self.request.user) | Q(pmDelegate=self.request.user)).
                                                                         values('project')). \
            values_list('id', 'project', 'project_id', 'project__projectId', 'project__name', 'template_id',
                        'lead_review_status'). \
            annotate(chapter_count=Count('project__book__chapter'))

        for k, v in header:
            worksheet.write(k, v)
        row = 1
        for s in project_details:
            worksheet.write(row, 0, s[3])#id
            worksheet.write(row, 1, s[4])#name
            worksheet.write(row, 2, self.request.user.get_full_name())#lead
            worksheet.write(row, 3, s[7])#chap count
            s1 = c_get_severity_count(s[1], "S1", s[5])
            s2 = c_get_severity_count(s[1], "S2", s[5])
            s3 = c_get_severity_count(s[1], "S3", s[5])
            worksheet.write(row, 4, s1)#s1
            worksheet.write(row, 5, s2)  # s2
            worksheet.write(row, 6, s3)  # s3
            q_count = c_get_question_count(s[1])
            worksheet.write(row, 7, q_count ) #q count
            worksheet.write(row, 8, c_get_defect_density(s1, s2, s3, q_count ))

            if c_get_project_status(s[1])[1] != 0:
                status = "Active"
            else:
                status = "Completed"
            worksheet.write(row, 9, status)
            row += 1

        workbook.close()

        return generateDownload(self.request, 'qms_project_summary.xlsx')


def review_completed(request):
    project_id = request.session['project']
    chapter_id = request.session['chapter']
    review_feedback = request.GET.get('review_feedback')
    review_group = request.GET.get('review_group')
    submitted_by = request.GET.get('submitted_by')
    try:
        if submitted_by == "author":
            QASheetHeader.objects.filter(project_id=project_id, chapter_component=request.session['chapter_component'],
                                         review_group_id=review_group).update(author_feedback_status=True,
                                                                              review_group_status=False,
                                                                              author_feedback=review_feedback)
        else:
            existing_remark = QASheetHeader.objects.filter(project_id=project_id,
                                                           chapter_component=request.session['chapter_component'],
                                                           review_group_id=review_group).values(
                "review_group_feedback", "author_feedback_status", "review_group_status").first()
            if len(str(existing_remark['review_group_feedback'])) > 0 and \
                    existing_remark['author_feedback_status'] == 1 and \
                    existing_remark['review_group_status'] == 0:

                review_feedback = "first review feedback :"+existing_remark['review_group_feedback']+"\n" +\
                                  "final review feedback : " + request.GET.get('review_feedback')
            QASheetHeader.objects.filter(project_id=project_id, chapter_component=request.session['chapter_component'],
                                         review_group_id=review_group).update(review_group_status=True,
                                                                              review_group_feedback=review_feedback)
            messages.success(request, "Saved Successfully")
    except Exception as e:
        messages.error(request, "Unable to save")
        logger.error("check permission for author failed {0} ".format(str(e)))
    obj = qa_sheet_header_obj(request.session['project'], request.session['chapter'], request.session['author'],
                              request.session['component'], review_group)

    qms_form = review_report_base(request.session['template_id'], request.session['project'],
                                  ChapterComponent.objects.get(chapter=request.session['chapter'],
                                  component=request.session['component']),
                                  request_obj=request, tab=obj.review_group_id)

    return render_common(obj, qms_form, request)


# https://en.wikipedia.org/wiki/Autovivification
def tree():
    return collections.defaultdict(tree)


def chapter_summary(request, p_id=None,export=False):
    if p_id is None:
        project_id = request.GET.get('project_id')
    else:
        project_id = p_id
    review_report_obj = ReviewReport.objects.filter(QA_sheet_header__project_id=project_id).\
        values('QA_sheet_header__chapter_id', 'QA_sheet_header__chapter_component_id').distinct().\
        annotate(cc_count=Count('QA_sheet_header__chapter_id', 'QA_sheet_header__chapter_component_id'))
    # below let us to assign without explicitly declaring index
    qms_data = tree()
    qms_data_list = []
    tmp_dict = {}
    severity_level = SeverityLevelMaster.objects.all()  # .exclude(name__icontains='S0')
    if review_report_obj:
        for eachData in review_report_obj:
            for k, v in eachData.iteritems():
                if k is 'QA_sheet_header__chapter_component_id':
                    try:
                        chapter_component_obj = ChapterComponent.objects.get(pk=v)
                        qms_data[chapter_component_obj.chapter.id][chapter_component_obj.component.id]['severity_level']\
                            = {}
                        qms_data[chapter_component_obj.chapter.id][chapter_component_obj.component.id]['chapter_name']\
                            = chapter_component_obj.chapter.name
                        qms_data[chapter_component_obj.chapter.id][chapter_component_obj.component.id]['component_name']\
                            = chapter_component_obj.component.name

                        tmp_obj = QASheetHeader.objects.filter(project_id=project_id,
                                                               chapter_component=chapter_component_obj)
                        question_count = tmp_obj.aggregate(Sum('count'))
                        question_count = question_count['count__sum']
                        qa_obj = tmp_obj[0]
                        qms_data[chapter_component_obj.chapter.id][chapter_component_obj.component.id]['author']\
                            = qa_obj.author.username
                        for s in severity_level:
                            s_count = review_report_obj.filter(defect_severity_level__severity_level=s,
                                                               QA_sheet_header__chapter_component=
                                                               chapter_component_obj). \
                                values('defect_severity_level__severity_level__name',
                                       'defect_severity_level__severity_level'). \
                                annotate(s_count=Count('defect_severity_level__severity_level')).exclude(
                                is_active=False)
                            if not s_count:
                                qms_data[chapter_component_obj.chapter.id][chapter_component_obj.component.id][
                                    'severity_level'][s.name] = 0
                                tmp_dict[s.name] = 0
                            else:
                                for c in s_count:
                                    try:
                                        qms_data[chapter_component_obj.chapter.id][chapter_component_obj.component.id][
                                            'severity_level'][s.name] = c['s_count']
                                        tmp_dict[s.name] = (c['s_count'] * s.penalty_count) / question_count
                                    except Exception as e:
                                        logger.error(" qms {0} ".format(str(e)))
                            tmp_dd = float(sum(tmp_dict.values()) * 100)
                            qms_data[chapter_component_obj.chapter.id][chapter_component_obj.component.id][
                                'defect_density'] = str(round(tmp_dd, 2))
                            qms_data[chapter_component_obj.chapter.id][chapter_component_obj.component.id][
                                'questions'] = question_count
                    except Exception as e:
                        # print "in except"
                        logger.error(" qms {0} ".format(str(e)))

                        # qms_data[obj.id]['severity_level'] = tmp_dict
            qms_data_list.append(qms_data.copy())
            qms_data.clear()
    if not export:
        return HttpResponse(
            json.dumps(qms_data_list),
            content_type="application/json"
        )
    else:
        return qms_data_list


def write_qms_chapter_level_summary(request):
    qms_data_list = chapter_summary(request, p_id=request.POST.get("project_id"), export=True)
    workbook = xlsxwriter.Workbook('qms_project_chapter_level_summary.xlsx')
    worksheet = workbook.add_worksheet()
    header = [' Chapter', "Component", "Author", 'S1', 'S2', 'S3', 'Questions',
              'Defect Density']
    header_length = len(header)
    header_column = list(string.ascii_uppercase)[:header_length]
    header_column = [s + "1" for s in header_column]
    header = zip(header_column, header)
    for k, v in header:
        worksheet.write(k, v)
    row = 1
    for l in qms_data_list:
        for k, v in l.iteritems():
            for sub_key, sub_value in v.iteritems():
                worksheet.write(row, 0, sub_value['chapter_name'])
                worksheet.write(row, 1, sub_value['component_name'])# id
                worksheet.write(row, 2, sub_value['author'])
                worksheet.write(row, 3, sub_value["severity_level"][u'S1'])
                worksheet.write(row, 4, sub_value["severity_level"][u'S2'])
                worksheet.write(row, 5, sub_value["severity_level"][u'S3'])
                worksheet.write(row, 6, sub_value['questions'])
                worksheet.write(row, 7, sub_value['defect_density'])

        row += 1

    workbook.close()

    return generateDownload(request, 'qms_project_chapter_level_summary.xlsx')


class ReviewListView(ListView):
    model = QASheetHeader
    template_name = 'review_list.html'

    def get_context_data(self, **kwargs):
        context = super(ReviewListView, self).get_context_data(**kwargs)
        self.request.session['is_pm'] = False
        if self.request.user.groups.filter(name='DeliveryManager').exists():
            self.request.session['is_pm'] = True
            context['review_list'] = QASheetHeader.objects.filter((Q(author=self.request.user)|
                                                                   Q(reviewed_by=self.request.user)
                                                                   | Q(project__in=ProjectDetail.objects.filter(
                                                                    Q(deliveryManager=self.request.user) |
                                                                    Q(pmDelegate=self.request.user)).values('project'))),
                                                                  project__endDate__gte=datetime.date.today(),
                                                                  project__closed=False,).values('id', 'project',
                                                                                                 'project_id',
                                                                                                 'project__projectId',
                                                                                                 'project__name',
                                                                                                 'order_number',
                                                                                                 'chapter_component',
                                                                                                 'chapter_component_id',
                                                                                                 'review_group_id',
                                                                                                 'review_group__name')

        else:
            context['review_list'] = QASheetHeader.objects.filter(Q(project__endDate__gte=datetime.date.today()) &
                                                                   (Q(author=self.request.user) |
                                                                    Q(reviewed_by=self.request.user)),
                                                                  project__closed=False, project__active=True, )\
                .values('id', 'project', 'project_id', 'project__projectId', 'project__name', 'order_number',
                        'chapter_component', 'chapter_component_id', 'review_group_id', 'review_group__name')
        context['review_list'] = context['review_list'].exclude((Q(review_group_status=True) &
                                                                Q(author_feedback_status=True)) |
                                                                Q(project__qms_project__lead_review_status=True))
        return context


class ReviewRedirectView(View):
    def get(self, request, *args, **kwargs):
        qa_obj = QASheetHeader.objects.get(id=self.kwargs['id'])
        request.session['QA_sheet_header_id'] = self.kwargs['id']
        self.request.session['chapter_component'] = cc_obj = ChapterComponent.objects.get(
            id=self.kwargs['chapter_component_id'])
        form = BaseAssessmentTemplateForm(initial={'project': qa_obj.project, 'chapter': cc_obj.chapter,
                                                   'author': qa_obj.author, 'component': cc_obj.component})
        self.request.session['active_tab'] = active_tab = self.kwargs['review_group_id']

        self.request.session['project'] = project = qa_obj.project
        self.request.session['chapter'] = chapter = cc_obj.chapter
        self.request.session['author'] = author = qa_obj.author
        self.request.session['component'] = component = cc_obj.component
        if self.request.user is author:
            get_review_group(project, chapter, is_author=False, component=component)
        try:
            ptpm_obj = ProjectTemplateProcessModel.objects.get(project=project)
            self.request.session['template_id'] = ptpm_obj.template_id
            self.request.session['producttype'] = QMSProcessModel.objects.get(name=ptpm_obj.qms_process_model).product_type
            obj = qa_sheet_header_obj(project, chapter, author, component, active_tab)
            is_pm = ProjectDetail.objects.filter(Q(deliveryManager=request.user) |
                                                  Q(pmDelegate=request.user), project=project).exists()
            self.request.session['is_pm'] = is_pm
            if obj.order_number != 1:
                order_number = int(obj.order_number) - 1
                prev_tab_obj = QASheetHeader.objects.filter(project=project,
                                                            chapter_component=self.request.session['chapter_component'],
                                                            order_number=order_number).first()
                if is_pm:
                    if prev_tab_obj.review_group_status is True and \
                                    prev_tab_obj.author_feedback_status is True:
                        pass
                    else:
                        return forbidden_access(self, form, project, "previous_tab_wait_pm", chapter)
                if not is_pm:
                    if prev_tab_obj.review_group_status and prev_tab_obj.author_feedback_status:
                        pass
                    else:
                        return forbidden_access(self, form, project, "previous_tab_wait", chapter)

            if self.request.user == obj.reviewed_by:
                self.request.session['reviewer_logged_in'] = True
                if obj.review_group_status is True and obj.author_feedback_status is False:
                    messages.info(self.request, " Please wait till author submit their feedback")
            else:
                self.request.session['reviewer_logged_in'] = False

            if obj is None:
                if is_pm:
                    msg = "config_missing_manager"
                else:
                    msg = "config_missing"
                return forbidden_access(self, form, project, msg, chapter)
            else:
                qms_team_members = [obj.reviewed_by, author]
                if not is_pm:
                    if self.request.user == author:
                        self.request.session['author_logged_in'] = True
                        if not obj.review_group_status and not obj.author_feedback_status:
                            return forbidden_access(self, form, project, "wait", chapter)
                    else:
                        self.request.session['author_logged_in'] = False
                else:
                    self.request.session['author_logged_in'] = False
                if not is_pm and request.user not in qms_team_members:
                    return forbidden_access(self, form, project, "not_assigned", chapter)

            project_template_process_model_obj = ProjectTemplateProcessModel.objects.get(project=project)
            template_id = request.session['template_id'] = project_template_process_model_obj.template_id
            qms_form = review_report_base(template_id, project, ChapterComponent.objects.get(chapter=chapter,
                                                                                             component=component),
                                          request_obj=self.request, tab=obj.review_group_id)

            return render_common(obj, qms_form, request)

        except:
            return render(request, "ansrS_QA_Tmplt_Assessment (Non Platform) QA sheet_3.3.html",
                          {'form': BaseAssessmentTemplateForm(), })


# s = ['Scripts_Review_dt',	'Copyedit_dt',	'ClientRound_dt',	'GDReview_dt',	'Beta_Round_1_dt',
#      'Beta_Round_2_dt',
#      'Client_Beta_dt',	'Gold_dt',	'Client_Gold_dt',	'Live_dt']


class ExportReview(View):
    def get(self, request, *args, **kwargs):
        review_obj = ReviewReport.objects.filter(QA_sheet_header_id=self.request.session['QA_sheet_header_id'],
                                                 is_active=True). \
            values_list('review_item', 'defect', 'defect_severity_level__severity_type__name',
                        'defect_severity_level__severity_level__name',
                        'defect_severity_level__defect_classification__name',
                        'is_fixed', 'fixed_by__username', 'remarks','instruction',).order_by('id')
        ptpm_obj = ProjectTemplateProcessModel.objects.get(project=QASheetHeader.objects.filter
                                                           (pk=self.request.session['QA_sheet_header_id']).values_list
                                                           ('project', flat=True).first())
        actual_name = ptpm_obj.template.actual_name
        print actual_name
        is_media = False
        if int(ptpm_obj.qms_process_model.product_type) == 1:
            is_media = True
            self.request.session['media'] = True
        else:
            self.request.session['media'] = False
        src = "QMS/master_templates/"+actual_name+".xlsx"
        if not os.path.isfile(src):
            src = "QMS/master_templates/QMS-T1.xlsx"

            logger.error(" failed to find template {0} ".format(actual_name))
        if self.request.session['active_tab'] == "lambda":
            review_group = review_obj.values_list('QA_sheet_header__review_group_id', flat=True).order_by('QA_sheet_header__order_number').first()
            self.request.session['active_tab'] = review_group
        else:
            review_group = self.request.session['active_tab']
        review_group = ReviewGroup.objects.get(id=review_group)
        if is_media:
            list_formula = review_group.review_master.name + u"_dt"
            # list_formula = 'INDIRECT'


        file_name = unicode(ptpm_obj.project)+" : " +self.request.session['chapter_component'].get_combination_name()\
                    + " : "+review_group.alias
        dst = "QMS/"+file_name+"_copy.xlsx"
        copyfile(src, dst)
        try:
            wb = load_workbook(filename=dst, data_only=False, guess_types=False)
            if is_media:
                ws1 = wb['Sheet1']
            else:
                ws1 = wb["Template1"]
            ws1.title = review_group.alias
            if not is_media:
                dv = DataValidation(type="list", formula1='Ref!$D$03:$D$100', allow_blank=True)
            else:
                dv = DataValidation(type="list", formula1="='Data Validation New'!$J$2:$J$11", allow_blank=True)
            ws1.add_data_validation(dv)

            c = 3
            s = 1
            for row in review_obj:
                c = str(c)
                # for i, header in enumerate(header)
                ws1["A" + c] = s
                ws1["B"+c].value = row[0]
                ws1["C"+c].value = row[1]
                ws1["D"+c].value = row[2]
                ws1["E" + c].value = row[3]
                ws1["F" + c].value = row[2]
                dv.add(ws1["D"+c])
                if not is_media:
                    severity_formula = DataValidation(type="custom", formula1='IFNA(VLOOKUP(D' + c + ',name,2,),"")')
                else:
                    severity_formula = DataValidation(type="custom", formula1='INDIRECT(INDIRECT("RC[-1]",0))')
                ws1.add_data_validation(severity_formula)
                ws1["E" + c].value = row[3]
                severity_formula.add(ws1["E" + c])
                if not is_media:
                    classification_formula = DataValidation(type="custom", formula1='IFNA(VLOOKUP(D' + c + ',name,3,),"")')
                else:
                    pass

                    classification_formula = DataValidation(type="custom", formula1='IF(D'+c + '<> "",'
                                                                                      'VLOOKUP($D'+c+', ''Defect_severity,'
                                                                                                        ' 2, FALSE), "")')

                ws1.add_data_validation(classification_formula)
                ws1["F" + c].value = row[4]
                classification_formula.add(ws1["E" + c])
                ws1["G"+c].value = row[5]
                ws1["H"+c].value = row[6]
                ws1["I"+c].value = row[7]
                ws1["J" + c].value = row[8]
                c = int(c)
                c += 1
                s += 1
            # dv.add(ws1)
            dv.ranges.append('D3:D500')
            wb.save(filename=dst)
        except Exception as e:
            logger.exception(e)

        return generateDownload(self.request, dst)


@transaction.atomic
def import_review(request, form_file):

    xl_workbook = xlrd.open_workbook(file_contents=form_file.read())

    sheet_names = xl_workbook.sheet_names()

    current_sheet = xl_workbook.sheet_by_name(sheet_names[0])
    classification_master = DefectClassificationMaster.objects.all().values_list('id','name')
    classification_master = dict((y.rstrip(), x) for x, y in classification_master)
    severity_level_master = SeverityLevelMaster.objects.all().values_list('id', 'name')
    severity_level_master = dict((y.rstrip(), x) for x, y in severity_level_master)
    defect_type_master = DefectTypeMaster.objects.all().values_list('id', 'name')
    defect_type_master = dict((y.rstrip(), x) for x, y in defect_type_master)
    ReviewReport.objects.filter(QA_sheet_header_id=request.session['QA_sheet_header_id']).update(is_active=False)

    l = current_sheet.nrows
    try:
        for r in range(2, l):
            if current_sheet.row(r)[3].value == "":
                break
            if current_sheet.row(r)[6].value != "":
                try:
                    fixed_by = User.objects.get(username=current_sheet.row(r)[7].value)
                except:
                    fixed_by = request.user
            else:
                fixed_by = None

            try:
                dtm = defect_type_master[(current_sheet.row(r)[3].value.rstrip())]
            except :
                dtm, created = DefectTypeMaster.objects.get_or_create(name=current_sheet.row(r)[3].value.rstrip(),
                                                                      defaults={'created_by': request.user})
                dtm = dtm.id

            try:
                if request.session['media']:
                    if request.session['media'] == True:
                        cm = classification_master[(current_sheet.row(r)[4].value.rstrip())]
                else:
                    cm = classification_master[(current_sheet.row(r)[5].value.rstrip())]

            except :
                if request.session['media']:
                    if request.session['media'] == True:
                        cm, created = DefectClassificationMaster.objects.get_or_create(name=current_sheet.row(r)[4].value.rstrip(),
                                                                               defaults={'created_by': request.user})
                        cm = cm.id
                else:
                    cm, created = DefectClassificationMaster.objects.get_or_create(name=current_sheet.row(r)[5].value.rstrip(), defaults={'created_by': request.user})
                    cm = cm.id

            try:
                if request.session['media']:
                    if request.session['media'] == True:
                        slm = severity_level_master[current_sheet.row(r)[5].value.rstrip()]
                else:
                    slm = severity_level_master[current_sheet.row(r)[4].value.rstrip()]

            except :
                if request.session['media']:
                    if request.session['media'] == True:
                        slm, created = SeverityLevelMaster.objects.get_or_create(name=current_sheet.row(r)[5].value.rstrip(),
                                                                                 penalty_count=0,
                                                                                 defaults={'created_by': request.user})
                        slm = slm.id
                else:
                    slm, created = SeverityLevelMaster.objects.get_or_create(name=current_sheet.row(r)[4].value.rstrip(),
                                                                             penalty_count=0,
                                                                             defaults={'created_by': request.user})
                    slm = slm.id
            if request.session['media']:
                if request.session['media'] == True:
                    dsl, created = DefectSeverityLevel.objects.get_or_create(severity_type_id=dtm, defect_classification_id=cm,
                                                                     severity_level_id=slm, product_type=1,
                                                                     defaults={'created_by': request.user})

                    ReviewReport.objects.create(QA_sheet_header_id=request.session['QA_sheet_header_id'],
                                        review_item=current_sheet.row(r)[1].value, defect=current_sheet.row(r)[2].value,
                                        defect_severity_level=dsl, is_fixed=current_sheet.row(r)[6].value,
                                        fixed_by=fixed_by, remarks=current_sheet.row(r)[8].value,
                                        instruction=current_sheet.row(r)[9].value,
                                        created_by=request.user)
                    qa_obj = QASheetHeader.objects.get(pk=request.session['QA_sheet_header_id'])
                    # messages.success(request, "successfully imported")
            else:
                dsl, created = DefectSeverityLevel.objects.get_or_create(severity_type_id=dtm, defect_classification_id=cm,
                                                                         severity_level_id=slm, product_type=0,
                                                                         defaults={'created_by': request.user})

                ReviewReport.objects.create(QA_sheet_header_id=request.session['QA_sheet_header_id'],
                                            review_item=current_sheet.row(r)[1].value, defect=current_sheet.row(r)[2].value,
                                            defect_severity_level=dsl, is_fixed=current_sheet.row(r)[6].value,
                                            fixed_by=fixed_by, remarks=current_sheet.row(r)[8].value,
                                            instruction=current_sheet.row(r)[9].value,
                                            created_by=request.user)
                qa_obj = QASheetHeader.objects.get(pk=request.session['QA_sheet_header_id'])
        messages.success(request, "successfully imported")
        return HttpResponseRedirect(
            reverse(u'review_redirect_view', kwargs={'id': request.session['QA_sheet_header_id'],
                                                     'chapter_component_id': qa_obj.chapter_component.id,
                                                     'review_group_id': qa_obj.review_group.id}))
    except Exception as e:
        messages.error(request, "please click on Export button")
        qa_obj = QASheetHeader.objects.get(pk=request.session['QA_sheet_header_id'])
        return HttpResponseRedirect(reverse(u'review_redirect_view', kwargs={'id': request.session['QA_sheet_header_id'],
                                        'chapter_component_id': qa_obj.chapter_component.id,
                                                                         'review_group_id': qa_obj.review_group.id}))

    # return HttpResponseRedirect(reverse_lazy(u'review_list'))